#!/usr/bin/python
# -*- coding: gb2312 -*-
#*******************************************************************************
# Copyright (c) 2012������ͨѶ�ɷ����޹�˾��All rights reserved.
# 
# �ļ����ƣ�tecs_common.py
# �������ݣ�
# ��ǰ�汾��1.0 
# ��    �ߣ���Т��
# ������ڣ�2012/11/27
#*******************************************************************************/
import sys,os,time

from tecs_common import *
from  openpyxl.reader.excel  import  load_workbook

from  openpyxl.workbook  import  Workbook   
#����ExcelWriter���õķ�װ���˲���˵����װ�˺�ǿ���excelд�Ĺ���    
from  openpyxl.writer.excel  import  ExcelWriter   
#һ��eggache������תΪ����ĸ�ķ���    
from  openpyxl.cell  import  get_column_letter  

class xlsx_sheet():
    def __init__(self, file_name):
        # �ȼ���ļ��Ƿ����
        try:
            wb = load_workbook(filename = file_name + '.xlsx')
        #print "load workbook success"
        except:    
            wb = Workbook()

        if wb == None:
            wb = Workbook()
        
        self.wb = wb
        # ����һ��excel��д����    
        ew = ExcelWriter(workbook = wb)
        self.ew = ew
        
        # �������·��
        dest_filename = file_name + '.xlsx'
        self.file_name = dest_filename
        

           
    def open(self, sheet_name):
        self.sheet_name = sheet_name
        # �����Ƿ����
        ws = self.wb.get_sheet_by_name(sheet_name.decode("gb2312"))
        self.ws = ws
        if ws == None:
            if self.wb.worksheets[0].title == "Sheet":
                ws = self.wb.worksheets[0]
            else:
                ws = self.wb.create_sheet()
            ws.title = sheet_name.decode("gb2312")
            
            ws.cell(row=0, column=0).value = "һ��������Ŀ".decode("gb2312")
            ws.cell(row=0, column=1).value = "����������Ŀ".decode("gb2312")
            ws.cell(row=0, column=2).value = "������ͼ".decode("gb2312")
            ws.cell(row=0, column=3).value = "��������".decode("gb2312")
            ws.cell(row=0, column=4).value = "�Ƿ�ʵ��".decode("gb2312")
            self.ws = ws
           
    def __del__(self):
        self.ew.save(filename = self.file_name)
        
    def write_xlsx_columu(self, column_n, write_num, record1="NULL", record2="NULL", recode3="NULL"):
        if write_num == 1:
            self.ws.cell(row=self.ws.get_highest_row(), column=column_n).value = record1.decode("gb2312")
        else:
            # ָ����д������
            write_row = self.ws.get_highest_row()
            self.ws.cell(row=write_row, column=column_n).value = record1.decode("gb2312")
            self.ws.cell(row=write_row, column=column_n+1).value = record2.decode("gb2312")
            self.ws.cell(row=write_row, column=column_n+2).value = recode3.decode("gb2312")

    # д��һ��������Ŀ����    
    def write_level_1(self, level1):    
        self.write_xlsx_columu(0, 1, level1)

    # д�����������Ŀ����     
    def write_level_2(self, level2):    
        self.write_xlsx_columu(1, 1, level2)    

    # д�������ͼ�Ͳ��Բ���
    def write_level_3(self, test_mean, test_step, is_do):    
        self.write_xlsx_columu(2, 3, test_mean, test_step, is_do)   

class cidoc():
    def __init__(self, file_name):
        # �Ȼ�ȡEXCEL�ľ��
        self.xlsx = xlsx_sheet(file_name)
        
    def write_doc(self, test_class, sheet_name):
        '''�Ѳ���������ӡ�����ȥ�����Դ����濴����ǰ��������Щ����'''
        self.xlsx.open(sheet_name)
        self.xlsx.write_level_1(test_class.__doc__.replace(" ", "").lstrip())
        num = 0
        for k,v in  test_class.__dict__.items():
            if type(getattr(test_class,k)) == type(test_class.__init__) and k.find("test_level_2") != -1:
                if getattr(test_class,k).__doc__ == None:
                    ci_exit(-1,"%s %s: find test function does not comment!" % (test_class, k))
                self.xlsx.write_level_2(getattr(test_class,k).__doc__.replace(" ", "").lstrip())
                level_2_name = str(k)
                level_3_name = level_2_name[:len(level_2_name)-1] + '3'
                for k1,v1 in  test_class.__dict__.items():
                    if type(getattr(test_class,k1)) == type(test_class.__init__) and k1.find(level_3_name) != -1:
                        # ��ȡע���ַ���
                        level_3_doc = getattr(test_class,k1).__doc__
                        if level_3_doc == None:
                            ci_exit(-1,"%s %s:find test function does not comment!" % (test_class,k1))
                        # ���ע�͵ĸ�ʽ�Ƿ��������
                        if level_3_doc.count("��") != 3:
                            ci_exit_ex(-1, line(), "%s %s: test note format error!" % (test_class,k1))
                        # ��ȡ�Ƿ�ʵ��
                        start_postion =  level_3_doc.find("��") + 2
                        end_postion =  level_3_doc.find("\n", start_postion)
                        is_do = level_3_doc[start_postion:end_postion].lstrip()
                        # ��ȡ������ͼ                    
                        start_postion =  level_3_doc.find("��", end_postion) + 2
                        end_postion =  level_3_doc.find("��", start_postion) - 8
                        test_mean = level_3_doc[start_postion:end_postion].replace(" ", "").lstrip()
                        # ��ȡ���Բ���
                        start_postion =  level_3_doc.find("��", end_postion) + 2
                        test_step = level_3_doc[start_postion:].replace(" ", "").lstrip()
                        # д��EXCEL
                        self.xlsx.write_level_3(test_mean, test_step, is_do) 
                        num += 1
                        print "write test itme doc %d" % num                    

        

        